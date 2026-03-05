"""
Excel Rapor Oluşturma Utility
Tarama sonuçlarını Excel dosyasına dönüştürür
"""
import logging
from typing import List, Dict
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

class ExcelGenerator:
    """Excel rapor oluşturur"""
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def generate(self, filepath: str, data: List[Dict]) -> str:
        """Basit veri listesinden Excel oluştur"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapor"
            
            # Başlık satırı
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                # Veri satırları
                for row_idx, item in enumerate(data, 2):
                    for col_idx, (key, value) in enumerate(item.items(), 1):
                        ws.cell(row=row_idx, column=col_idx).value = value
            
            # Dosyayı kaydet
            Path(filepath).parent.mkdir(exist_ok=True)
            wb.save(filepath)
            logger.info(f"✅ Excel raporu oluşturuldu: {filepath}")
            
            return filepath
        except Exception as e:
            logger.error(f"❌ Excel raporu oluşturma hatası: {e}")
            raise
    
    def generate_report(self, products_data: List[Dict], store_name: str = "Trendyol") -> str:
        """Rapor oluştur ve dosya yolunu döndür"""
        try:
            # Workbook oluştur
            wb = openpyxl.Workbook()
            
            # Ana rapor sayfası
            ws = wb.active
            ws.title = "Tarama Raporu"
            
            # Başlık ekle
            self._add_header(ws, store_name)
            
            # Veri ekle
            self._add_data(ws, products_data)
            
            # Otomatik filtre ekle
            self._add_filters(ws)
            
            # Koşullu biçimlendirme ekle
            self._add_conditional_formatting(ws, products_data)
            
            # Özet sayfası ekle (Ürün başına en ucuz satıcı)
            self._add_summary_sheet(wb, products_data)
            
            # Dosyayı kaydet
            filename = "trendyol_rapor.xlsx"
            filepath = self.output_dir / filename
            
            wb.save(str(filepath))
            logger.info(f"✅ Excel raporu oluşturuldu: {filepath}")
            logger.info(f"   📊 Sayfalar: Tarama Raporu, Özet Analiz")
            
            return str(filepath)
        except Exception as e:
            logger.error(f"❌ Excel raporu oluşturma hatası: {e}")
            raise
    
    def _add_header(self, ws, store_name: str) -> None:
        """Başlık ekle"""
        # Başlık satırı
        ws['A1'] = f"Trendyol Satıcı Analiz Raporu - {store_name}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:I1')
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Tarih bilgisi
        ws['A2'] = f"Rapor Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        # Sütun başlıkları
        headers = [
            "Ürün Adı",
            "Ürün Linki",
            "Satıcı",
            "Orijinal Fiyat (TL)",
            "Kupon İndirimi",
            "Sepette İndirimi",
            "Son Fiyat (TL)",
            "Rating",
            "Notlar"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def _add_data(self, ws, products_data: List[Dict]) -> None:
        """Veri ekle"""
        row = 5
        
        for product in products_data:
            product_name = product.get('name', '')
            sellers = product.get('sellers', [])
            
            if not sellers:
                continue
            
            # Her satıcı için bir satır
            for idx, seller in enumerate(sellers):
                # Ürün adı (HER satıcı satırında)
                # Satıcıdan çekilmiş doğru adı kullan, yoksa varsayılan adı kullan
                actual_name = seller.get('product_name', product_name)
                ws.cell(row=row, column=1).value = actual_name
                
                # Ürün Linki
                ws.cell(row=row, column=2).value = product.get('url', '')
                
                # Satıcı adı
                ws.cell(row=row, column=3).value = seller.get('name', '')
                
                # Orijinal fiyat
                ws.cell(row=row, column=4).value = seller.get('price', 0)
                
                # Kupon İndirimi
                ws.cell(row=row, column=5).value = seller.get('coupon', '')
                
                # Sepette İndirimi
                ws.cell(row=row, column=6).value = seller.get('basket_discount', '')
                
                # Son Fiyat (Net Fiyat)
                ws.cell(row=row, column=7).value = seller.get('net_price', 0)
                
                # Rating
                ws.cell(row=row, column=8).value = seller.get('rating', 0)
                
                # Notlar
                notes = []
                if seller.get('coupon'):
                    notes.append(f"Kupon: {seller.get('coupon')}")
                if seller.get('basket_discount'):
                    notes.append(f"Sepette: {seller.get('basket_discount')}")
                ws.cell(row=row, column=9).value = ' | '.join(notes)
                
                # Hücre formatı
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Sayı formatı
                    if col in [4, 7, 8]:
                        cell.number_format = '0.00'
                
                row += 1
        
        # Sütun genişliği ayarla
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 30
    
    def _add_filters(self, ws) -> None:
        """Otomatik filtre ekle"""
        try:
            # Başlık satırına filtre ekle (satır 4)
            ws.auto_filter.ref = f"A4:I{ws.max_row}"
            logger.info("   ✅ Otomatik filtre eklendi")
        except Exception as e:
            logger.warning(f"⚠️ Filtre ekleme hatası: {e}")
    
    def _add_conditional_formatting(self, ws, products_data: List[Dict]) -> None:
        """Koşullu biçimlendirme ekle - En ucuz fiyatı yeşil, en pahalı fiyatı kırmızı"""
        try:
            from openpyxl.formatting.rule import CellIsRule
            
            # Ürün başına en ucuz ve en pahalı fiyatları bul
            product_prices = {}
            for product in products_data:
                product_name = product.get('name', '')
                sellers = product.get('sellers', [])
                
                prices = []
                for seller in sellers:
                    net_price = seller.get('net_price', 0)
                    if net_price > 0:
                        prices.append(net_price)
                
                if prices:
                    product_prices[product_name] = {
                        'min': min(prices),
                        'max': max(prices)
                    }
            
            # Fiyat sütununda (G sütunu) koşullu biçimlendirme uygula
            row = 5
            for product in products_data:
                product_name = product.get('name', '')
                sellers = product.get('sellers', [])
                
                if product_name in product_prices:
                    min_price = product_prices[product_name]['min']
                    max_price = product_prices[product_name]['max']
                    
                    for seller in sellers:
                        net_price = seller.get('net_price', 0)
                        
                        # En ucuz fiyat - yeşil
                        if net_price == min_price and min_price > 0:
                            cell = ws.cell(row=row, column=7)
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            cell.font = Font(bold=True, color="006100")
                        
                        # En pahalı fiyat - kırmızı
                        elif net_price == max_price and max_price > 0:
                            cell = ws.cell(row=row, column=7)
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            cell.font = Font(bold=True, color="9C0006")
                        
                        row += 1
                else:
                    row += len(sellers)
            
            logger.info("   ✅ Koşullu biçimlendirme eklendi (En ucuz: yeşil, En pahalı: kırmızı)")
        except Exception as e:
            logger.warning(f"⚠️ Koşullu biçimlendirme hatası: {e}")
    
    def _add_summary_sheet(self, wb, products_data: List[Dict]) -> None:
        """Özet sayfası ekle - Ürün başına en ucuz satıcı"""
        try:
            ws_summary = wb.create_sheet("Özet Analiz")
            
            # Başlık
            ws_summary['A1'] = "Ürün Başına En Ucuz Satıcı Analizi"
            ws_summary['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws_summary.merge_cells('A1:F1')
            ws_summary['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Sütun başlıkları
            headers = [
                "Ürün Adı",
                "En Ucuz Satıcı",
                "En Ucuz Fiyat (TL)",
                "En Pahalı Fiyat (TL)",
                "Fiyat Farkı (TL)",
                "Fiyat Farkı (%)"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Veri ekle
            row = 4
            for product in products_data:
                product_name = product.get('name', '')
                sellers = product.get('sellers', [])
                
                if not sellers:
                    continue
                
                # Satıcı bilgilerini topla
                seller_info = []
                for seller in sellers:
                    net_price = seller.get('net_price', 0)
                    if net_price > 0:
                        seller_info.append({
                            'name': seller.get('name', ''),
                            'price': net_price,
                            'rating': seller.get('rating', 0)
                        })
                
                if not seller_info:
                    continue
                
                # En ucuz ve en pahalı fiyatları bul
                seller_info_sorted = sorted(seller_info, key=lambda x: x['price'])
                min_seller = seller_info_sorted[0]
                max_seller = seller_info_sorted[-1]
                
                min_price = min_seller['price']
                max_price = max_seller['price']
                price_diff = max_price - min_price
                price_diff_percent = (price_diff / min_price * 100) if min_price > 0 else 0
                
                # Ürün adı
                ws_summary.cell(row=row, column=1).value = product_name
                
                # En ucuz satıcı
                ws_summary.cell(row=row, column=2).value = f"{min_seller['name']} ({min_seller['rating']}⭐)"
                
                # En ucuz fiyat
                cell = ws_summary.cell(row=row, column=3)
                cell.value = min_price
                cell.number_format = '0.00'
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.font = Font(bold=True, color="006100")
                
                # En pahalı fiyat
                cell = ws_summary.cell(row=row, column=4)
                cell.value = max_price
                cell.number_format = '0.00'
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True, color="9C0006")
                
                # Fiyat farkı
                cell = ws_summary.cell(row=row, column=5)
                cell.value = price_diff
                cell.number_format = '0.00'
                
                # Fiyat farkı %
                cell = ws_summary.cell(row=row, column=6)
                cell.value = price_diff_percent
                cell.number_format = '0.00'
                
                # Hücre formatı
                for col in range(1, 7):
                    cell = ws_summary.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                row += 1
            
            # Sütun genişliği
            ws_summary.column_dimensions['A'].width = 35
            ws_summary.column_dimensions['B'].width = 25
            ws_summary.column_dimensions['C'].width = 18
            ws_summary.column_dimensions['D'].width = 18
            ws_summary.column_dimensions['E'].width = 18
            ws_summary.column_dimensions['F'].width = 18
            
            logger.info("   ✅ Özet Analiz sayfası eklendi")
        except Exception as e:
            logger.warning(f"⚠️ Özet sayfası oluşturma hatası: {e}")
